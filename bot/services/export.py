from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from bot.db import Video

APP_TIMEZONE = ZoneInfo("Europe/Kyiv")
HEADER_FILL = PatternFill("solid", fgColor="172A3A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LIGHT_BORDER = Border(bottom=Side(style="thin", color="D9E2F3"))
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
DETAIL_HEADERS = (
    "Пользователь",
    "Telegram ID",
    "Название",
    "Платформа",
    "Ссылка",
    "Дата добавления",
    "Статус",
    "Просмотры",
    "Лайки",
    "Комментарии",
    "Репосты",
    "Сумма оплаты, ₽",
    "Способ оплаты",
    "Реквизиты",
    "Статистика обновлена",
    "Результат обновления",
)
STATUS_LABELS = {
    "confirmed": "Подтверждён",
    "approved": "Одобрен",
    "paid": "Оплачен",
}
PLATFORM_LABELS = {
    "youtube": "YouTube Shorts",
    "tiktok": "TikTok",
    "instagram": "Instagram Reels",
}
PAYMENT_METHOD_LABELS = {
    "card": "Банковская карта",
    "usdt": "USDT TRC-20",
    "ggstore": "Баланс gg.store",
}


def create_videos_workbook(
    videos: list[Video],
    refresh_statuses: dict[int, str],
    output_path: Path,
) -> Path:
    """Создаёт проверенный Excel-отчёт по роликам."""

    workbook = Workbook()
    detail_sheet = workbook.active
    detail_sheet.title = "Ролики"
    _populate_detail_sheet(detail_sheet, videos, refresh_statuses)
    summary_sheet = workbook.create_sheet("Сводка", 0)
    _populate_summary_sheet(summary_sheet, videos)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    _verify_workbook(output_path, len(videos))
    return output_path


def build_export_filename(start_date: date, end_date: date) -> str:
    """Возвращает стабильное имя Excel-выгрузки."""

    return f"ugc_videos_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"


def _populate_detail_sheet(sheet, videos: list[Video], statuses: dict[int, str]) -> None:
    """Заполняет детальный лист роликов."""

    sheet.append(DETAIL_HEADERS)
    for video in videos:
        sheet.append(_build_detail_row(video, statuses.get(video.video_id, "Из базы")))
    _style_header(sheet, len(DETAIL_HEADERS))
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    _configure_print_layout(sheet)
    sheet.auto_filter.ref = f"A1:P{max(sheet.max_row, 1)}"
    if videos:
        table = Table(displayName="VideosTable", ref=f"A1:P{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    _format_detail_columns(sheet)


def _build_detail_row(video: Video, refresh_status: str) -> tuple[object, ...]:
    """Собирает типизированную строку одного ролика."""

    username = f"@{video.user.username}" if video.user and video.user.username else ""
    created_at = _local_naive(video.created_at)
    updated_at = _local_naive(video.views_updated_at) if video.views_updated_at else None
    return (
        username,
        video.user_id,
        video.title or "",
        PLATFORM_LABELS.get(video.platform, video.platform),
        video.url,
        created_at,
        STATUS_LABELS.get(video.status, video.status),
        video.views_count,
        video.likes_count,
        video.comments_count,
        video.shares_count,
        video.payout_amount,
        PAYMENT_METHOD_LABELS.get(
            video.user.payment_method if video.user else None,
            "Не указан",
        ),
        video.user.payment_details if video.user and video.user.payment_details else "",
        updated_at,
        refresh_status,
    )


def _populate_summary_sheet(sheet, videos: list[Video]) -> None:
    """Создаёт формульную сводку по пользователям."""

    headers = (
        "Пользователь",
        "Telegram ID",
        "Роликов",
        "Просмотры",
        "Лайки",
        "Комментарии",
        "Репосты",
        "Сумма оплаты, ₽",
    )
    sheet.append(headers)
    users: dict[int, str] = {}
    for video in videos:
        users[video.user_id] = f"@{video.user.username}" if video.user and video.user.username else ""
    detail_end = max(len(videos) + 1, 2)
    for row_index, (user_id, username) in enumerate(sorted(users.items()), start=2):
        sheet.append((username, user_id))
        sheet.cell(row_index, 3, f'=COUNTIF(\'Ролики\'!$B$2:$B${detail_end},B{row_index})')
        for column_index, detail_column in enumerate(("H", "I", "J", "K", "L"), start=4):
            formula = (
                f'=SUMIF(\'Ролики\'!$B$2:$B${detail_end},B{row_index},'
                f"'Ролики'!${detail_column}$2:${detail_column}${detail_end})"
            )
            sheet.cell(row_index, column_index, formula)
    total_row = sheet.max_row + 1
    sheet.cell(total_row, 1, "ИТОГО")
    for column in range(3, 9):
        letter = get_column_letter(column)
        sheet.cell(total_row, column, f"=SUM({letter}2:{letter}{total_row - 1})")
    for cell in sheet[total_row]:
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True)
        cell.border = LIGHT_BORDER
    _style_header(sheet, len(headers))
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = f"A1:H{max(sheet.max_row, 1)}"
    _configure_print_layout(sheet)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 16
    for column in range(3, 9):
        sheet.column_dimensions[get_column_letter(column)].width = 18
        for cell in sheet[get_column_letter(column)][1:]:
            cell.number_format = '#,##0.00' if column == 8 else '#,##0'
            cell.alignment = Alignment(horizontal="right")


def _style_header(sheet, column_count: int) -> None:
    """Оформляет строку заголовков."""

    for cell in sheet[1][:column_count]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 34


def _configure_print_layout(sheet) -> None:
    """Подготавливает широкую таблицу к печати на одной странице по ширине."""

    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = "1:1"
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.5
    sheet.page_margins.bottom = 0.5


def _format_detail_columns(sheet) -> None:
    """Применяет ширины, форматы и гиперссылки детального листа."""

    widths = (24, 16, 34, 20, 44, 20, 18, 16, 14, 16, 14, 18, 24, 34, 22, 24)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        row[4].hyperlink = row[4].value
        row[4].style = "Hyperlink"
        row[5].number_format = "yyyy-mm-dd hh:mm"
        row[14].number_format = "yyyy-mm-dd hh:mm"
        for index in range(7, 12):
            row[index].number_format = '#,##0'
            row[index].alignment = Alignment(horizontal="right")
        row[11].number_format = '#,##0.00'
        for cell in row:
            cell.border = LIGHT_BORDER
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical="top",
                wrap_text=False,
            )


def _local_naive(value: datetime) -> datetime:
    """Приводит datetime к часовому поясу отчёта и убирает tzinfo для Excel."""

    if value.tzinfo is None:
        return value
    return value.astimezone(APP_TIMEZONE).replace(tzinfo=None)


def _verify_workbook(path: Path, expected_rows: int) -> None:
    """Переоткрывает файл и проверяет ключевую структуру."""

    workbook = load_workbook(path, read_only=False, data_only=False)
    if workbook.sheetnames != ["Сводка", "Ролики"]:
        raise ValueError("Unexpected workbook sheets")
    if workbook["Ролики"].max_row != expected_rows + 1:
        raise ValueError("Unexpected workbook row count")
    workbook.close()
