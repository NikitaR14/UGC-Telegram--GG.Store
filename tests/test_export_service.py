from __future__ import annotations

from datetime import UTC, date, datetime
from types import SimpleNamespace

from openpyxl import load_workbook

from bot.handlers.admin.export import build_utc_range, parse_date
from bot.services.export import build_export_filename, create_videos_workbook


def build_video(video_id: int, username: str, views_count: int) -> SimpleNamespace:
    """Создаёт тестовую строку выгрузки."""

    user = SimpleNamespace(
        username=username,
        payment_method="card",
        payment_details="4111111111111234",
    )
    return SimpleNamespace(
        video_id=video_id,
        user_id=100 + video_id,
        user=user,
        title=f"Ролик {video_id}",
        platform="instagram",
        url=f"https://instagram.com/reel/{video_id}",
        created_at=datetime(2026, 8, video_id, 9, 30, tzinfo=UTC),
        status="confirmed",
        views_count=views_count,
        likes_count=500,
        comments_count=25,
        shares_count=10,
        payout_amount=25000.0,
        views_updated_at=datetime(2026, 8, 31, 12, 0, tzinfo=UTC),
    )


def test_create_videos_workbook_builds_detail_and_formula_summary(tmp_path) -> None:
    """Проверяет структуру, типы данных и формулы Excel-отчёта."""

    path = tmp_path / "videos.xlsx"
    videos = [build_video(1, "first", 120_000), build_video(2, "second", 200_000)]

    create_videos_workbook(videos, {1: "Обновлено", 2: "Из базы"}, path)
    workbook = load_workbook(path, data_only=False)
    detail = workbook["Ролики"]
    summary = workbook["Сводка"]

    assert workbook.sheetnames == ["Сводка", "Ролики"]
    assert detail.max_row == 3
    assert detail["H2"].value == 120_000
    assert isinstance(detail["F2"].value, datetime)
    assert detail["N2"].value == "4111111111111234"
    assert detail["M2"].value == "Банковская карта"
    assert detail["E2"].hyperlink.target == "https://instagram.com/reel/1"
    assert str(summary["C2"].value).startswith("=COUNTIF")
    assert str(summary["H2"].value).startswith("=SUMIF")
    assert summary["A4"].value == "ИТОГО"
    assert str(summary["H4"].value).startswith("=SUM")
    assert detail.page_setup.orientation == "landscape"
    assert detail.page_setup.fitToWidth == 1
    workbook.close()


def test_export_date_range_is_inclusive_in_kyiv_timezone() -> None:
    """Проверяет включительные границы локальных дат."""

    created_from, created_to = build_utc_range(date(2026, 8, 1), date(2026, 8, 31))

    assert created_from == datetime(2026, 7, 31, 21, 0, tzinfo=UTC)
    assert created_to == datetime(2026, 8, 31, 21, 0, tzinfo=UTC)
    assert parse_date("31.08.2026") == date(2026, 8, 31)
    assert parse_date("2026-08-31") is None
    assert build_export_filename(date(2026, 8, 1), date(2026, 8, 31)).endswith(".xlsx")
